"""Report generation — CSV and Excel export."""
from __future__ import annotations
import csv
import os
from datetime import date
from typing import Dict, List, Optional

from ..data.db_manager import DBManager
from ..utils.helpers import ensure_dir, get_app_data_dir
from ..utils.logger import get_logger

log = get_logger()

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    _OPENPYXL = True
except ImportError:
    _OPENPYXL = False


class ReportService:
    """Generates CSV and Excel reports."""

    def __init__(self, db: DBManager) -> None:
        self.db = db

    # ------------------------------------------------------------------ #
    # Data queries
    # ------------------------------------------------------------------ #

    def loans_summary(self, status: Optional[str] = None,
                      from_date: Optional[str] = None,
                      to_date: Optional[str] = None) -> List[Dict]:
        conditions = []
        params: list = []
        if status:
            conditions.append("l.status=?")
            params.append(status)
        if from_date:
            conditions.append("l.application_date>=?")
            params.append(from_date)
        if to_date:
            conditions.append("l.application_date<=?")
            params.append(to_date)
        where = ("WHERE " + " AND ".join(conditions)) if conditions else ""
        return self.db.fetch_all(
            f"SELECT l.reference_number, e.name AS employee, e.employee_code, "
            f"l.requested_amount, l.approved_amount, l.interest_rate, l.term_months, "
            f"l.status, l.application_date, l.due_date, l.outstanding_balance, "
            f"b.name AS branch "
            f"FROM loans l "
            f"LEFT JOIN employees e ON l.employee_id=e.id "
            f"LEFT JOIN branches b ON l.branch_id=b.id "
            f"{where} ORDER BY l.application_date DESC",
            tuple(params),
        )

    def repayments_summary(self, from_date: Optional[str] = None,
                           to_date: Optional[str] = None) -> List[Dict]:
        conditions = []
        params: list = []
        if from_date:
            conditions.append("r.payment_date>=?")
            params.append(from_date)
        if to_date:
            conditions.append("r.payment_date<=?")
            params.append(to_date)
        where = ("WHERE " + " AND ".join(conditions)) if conditions else ""
        return self.db.fetch_all(
            f"SELECT r.payment_date, l.reference_number, e.name AS employee, "
            f"r.amount, r.payment_method, r.reference, r.notes "
            f"FROM repayments r "
            f"LEFT JOIN loans l ON r.loan_id=l.id "
            f"LEFT JOIN employees e ON l.employee_id=e.id "
            f"{where} ORDER BY r.payment_date DESC",
            tuple(params),
        )

    def outstanding_balances(self) -> List[Dict]:
        return self.db.fetch_all(
            "SELECT l.reference_number, e.name AS employee, e.employee_code, "
            "l.approved_amount, l.outstanding_balance, l.due_date, l.status, "
            "b.name AS branch "
            "FROM loans l "
            "LEFT JOIN employees e ON l.employee_id=e.id "
            "LEFT JOIN branches b ON l.branch_id=b.id "
            "WHERE l.status IN ('active','approved') "
            "ORDER BY l.outstanding_balance DESC"
        )

    # ------------------------------------------------------------------ #
    # CSV export
    # ------------------------------------------------------------------ #

    def export_csv(self, rows: List[Dict], filename: str) -> str:
        """Write *rows* to *filename* as CSV. Returns the file path."""
        ensure_dir(os.path.dirname(filename))
        if not rows:
            return filename
        with open(filename, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=rows[0].keys())
            writer.writeheader()
            writer.writerows(rows)
        log.info("Exported CSV: %s (%d rows)", filename, len(rows))
        return filename

    # ------------------------------------------------------------------ #
    # Excel export
    # ------------------------------------------------------------------ #

    def export_excel(self, rows: List[Dict], filename: str,
                     sheet_name: str = "Report") -> str:
        """Write *rows* to *filename* as .xlsx. Returns the file path."""
        if not _OPENPYXL:
            log.warning("openpyxl not installed; falling back to CSV")
            return self.export_csv(rows, filename.replace(".xlsx", ".csv"))

        ensure_dir(os.path.dirname(filename))
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name

        if not rows:
            wb.save(filename)
            return filename

        headers = list(rows[0].keys())
        header_fill = PatternFill("solid", fgColor="2E86C1")
        header_font = Font(bold=True, color="FFFFFF")

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header.replace("_", " ").title())
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for row_idx, row in enumerate(rows, 2):
            for col_idx, key in enumerate(headers, 1):
                ws.cell(row=row_idx, column=col_idx, value=row.get(key))

        # Auto-size columns
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        wb.save(filename)
        log.info("Exported Excel: %s (%d rows)", filename, len(rows))
        return filename

    def default_export_path(self, report_type: str, ext: str = "xlsx") -> str:
        today = date.today().strftime("%Y%m%d")
        export_dir = os.path.join(get_app_data_dir(), "exports")
        ensure_dir(export_dir)
        return os.path.join(export_dir, f"{report_type}_{today}.{ext}")
